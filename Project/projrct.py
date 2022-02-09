import openpyxl


df2 = openpyxl.load_workbook("OfficeM&A_RootWords.xlsx")
sheet1 = df2['RootWords']
row = sheet1.max_row
column = sheet1.max_column
x3 = sheet1.cell(row=3, column=1)
#print(x3.value)
for i in range(1, row+1):
    for j in range(1, column):
        print(sheet1.cell(i, j).value)
