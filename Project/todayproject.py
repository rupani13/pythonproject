# We are printing a preticular column and the particular keyword

import openpyxl
import pandas as pd

df = pd.read_excel("OfficeM&A_RootWords.xlsx")
df1 = pd.read_excel("OfficeM&A_RootWords.xlsx", sheet_name='RootWords', usecols="B")
print(df1)

df1 = openpyxl.load_workbook("OfficeM&A_RootWords.xlsx")
print(type(df1))
sheets = df1.sheetnames
print(sheets)
sheet = df1['RootWords']
s = sheet['A1'].value
print(sheet.cell(1, 1).value)