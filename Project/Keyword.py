import openpyxl

df1 = openpyxl.load_workbook("OfficeM&A_RootWords.xlsx")

sheet = df1['RootWords']
r = sheet.max_row
c = sheet.max_column
obj = input("Enter the Keyword")
for i in range(1, r):
    for j in range(1, c):
        if sheet.cell(i, j).value != 1:
            print("obj not found")
        else:
            print("obj found")